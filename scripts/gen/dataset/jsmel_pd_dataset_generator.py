#!/usr/bin/env python3
"""
jsmel_pd_dataset_generator.py

Generate a neutral monophonic WAV corpus from the OPEN public-domain subset of
JSMel (547 melodies).

JSMel
-----
R. Matsunaga, T. Ishimoto, P. Hartono, J. Abe,
"JSMel: A dataset of 960 song melodies widely shared in contemporary Japan",
Data in Brief 66 (2026), 112869.
https://doi.org/10.1016/j.dib.2026.112869

Open PD repository:
https://doi.org/10.5281/zenodo.20078158

The public JSMel release contains symbolic melody data only for 547
public-domain compositions. Each melody is represented by:

    P_<ID>.txt  : pitch sequence, e.g. C4 D#4 rest C5
    D_<ID>.txt  : duration sequence, quarter-note = 1

The two sequences have equal length and matching indices.

Why this generator is useful for Scaleify
-----------------------------------------
Unlike an audio/MIDI corpus, JSMel has already:
- isolated the monophonic main melody,
- removed lyrics/accompaniment,
- removed ornamental notes,
- removed tempo/dynamics/expression/articulation,
- retained only the first verse when multiple verses exist.

Therefore no F0 estimation or melody-track heuristic is needed here:

    JSMel P/D sequences
        -> exact symbolic note/rest events
        -> one fixed neutral synth
        -> dataset/japan_jsmel/*.wav

The generated WAVs are compatibility files for the current Scaleify trainer.
The symbolic sequences remain the higher-quality source representation.

License / scope
---------------
The JSMel paper states that the openly released PD symbolic melody data and
metadata are available under CC BY-NC 4.0. The underlying 547 compositions are
classified by JSMel as public domain. Attribute JSMel and respect the
non-commercial restriction on the released dataset.

JSMel represents melodies widely familiar in contemporary Japan. It is NOT
limited to Japanese-origin traditional music; the complete 960-song collection
contains foreign-origin songs as well. Use --exclude-foreign-origin if you want
to conservatively remove records marked as foreign-origin in the metadata.

Dependencies
------------
numpy
requests
soundfile
openpyxl

Usage
-----
Download the latest open JSMel PD record and generate all available PD songs:

    python jsmel_pd_dataset_generator.py

List detected melody IDs without rendering:

    python jsmel_pd_dataset_generator.py --list

Generate one melody:

    python jsmel_pd_dataset_generator.py --song A001

Use an already downloaded/extracted JSMel directory:

    python jsmel_pd_dataset_generator.py --source-dir /path/to/jsmel_pd

Exclude metadata rows marked as songs of foreign origin:

    python jsmel_pd_dataset_generator.py --exclude-foreign-origin

Force re-download / re-render:

    python jsmel_pd_dataset_generator.py --force

Default output
--------------
dataset/japan_jsmel/
    A001.wav
    ...
    metadata.csv
    manifest.json
    _jsmel_downloads/
    _jsmel_source/
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import shutil
import tempfile
import time
import unicodedata
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
import requests
import soundfile as sf
from openpyxl import load_workbook


ZENODO_RECORD_ID = "20078158"
ZENODO_RECORD_API = f"https://zenodo.org/api/records/{ZENODO_RECORD_ID}"
ZENODO_LANDING = f"https://zenodo.org/records/{ZENODO_RECORD_ID}"
JSMEL_DOI = "10.5281/zenodo.20078158"
PAPER_DOI = "10.1016/j.dib.2026.112869"

DEFAULT_OUTPUT = Path("datasets/japan_jsmel")
DEFAULT_SR = 44100
DEFAULT_BPM = 100.0
DEFAULT_GAP_MS = 18.0

PITCH_RE = re.compile(
    r"^(?P<name>[A-Ga-g])"
    r"(?P<acc>(?:#{1,2}|b{1,2}|x|♯|♭)?)"
    r"(?P<oct>-?\d+)$"
)

NUMBER_RE = re.compile(
    r"[-+]?\d+/\d+|[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][-+]?\d+)?"
)

FALSEISH = {
    "", "0", "no", "n", "false", "none", "na", "n/a", "-", "not applicable",
    "japan", "japanese", "domestic",
}


@dataclass
class Event:
    pitch: int | None
    duration_q: float


@dataclass
class Pair:
    melody_id: str
    pitch_path: Path
    duration_path: Path


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def safe_filename(name: str) -> str:
    name = Path(name).name
    if not name or name in {".", ".."}:
        raise ValueError(f"Unsafe filename: {name!r}")
    return name


def fetch_json(session: requests.Session, url: str, timeout: float) -> dict:
    r = session.get(url, timeout=timeout)
    r.raise_for_status()
    return r.json()


def download_file(
    session: requests.Session,
    url: str,
    path: Path,
    timeout: float,
    force: bool,
) -> None:
    if path.exists() and not force:
        return

    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".part")

    with session.get(url, timeout=timeout, stream=True) as r:
        r.raise_for_status()
        with tmp.open("wb") as f:
            for chunk in r.iter_content(chunk_size=1024 * 1024):
                if chunk:
                    f.write(chunk)

    tmp.replace(path)


def record_files(record: dict) -> list[dict]:
    files = record.get("files") or []
    out = []

    for item in files:
        key = item.get("key") or item.get("filename")
        links = item.get("links") or {}
        url = links.get("content") or links.get("self")
        if not key or not url:
            continue
        out.append({
            "key": safe_filename(str(key)),
            "url": str(url),
            "size": item.get("size"),
            "checksum": item.get("checksum"),
        })

    return out


def acquire_jsmel(
    output: Path,
    timeout: float,
    force: bool,
    request_delay: float,
) -> tuple[Path, Path | None, dict]:
    """
    Download all ZIPs from the current Zenodo PD record plus Metadata_JSMel.xlsx,
    then safely extract the archives.
    """
    downloads = output / "_jsmel_downloads"
    extracted = output / "_jsmel_source"
    downloads.mkdir(parents=True, exist_ok=True)

    session = requests.Session()
    session.headers.update({
        "User-Agent": "Scaleify-JSMel-PD-generator/1.0 (research use)"
    })

    record = fetch_json(session, ZENODO_RECORD_API, timeout)
    files = record_files(record)

    zip_entries = [
        f for f in files
        if f["key"].lower().endswith(".zip")
    ]
    metadata_entries = [
        f for f in files
        if f["key"].lower().endswith((".xlsx", ".xlsm"))
        and "metadata" in f["key"].lower()
    ]

    if not zip_entries:
        raise RuntimeError(
            "Zenodo record returned no ZIP files. "
            "The record structure may have changed."
        )

    downloaded_zips: list[Path] = []
    for i, item in enumerate(sorted(zip_entries, key=lambda x: x["key"])):
        local = downloads / item["key"]
        print(f"Downloading {item['key']} ...")
        download_file(session, item["url"], local, timeout, force)
        downloaded_zips.append(local)
        if request_delay > 0 and i + 1 < len(zip_entries):
            time.sleep(request_delay)

    metadata_path = None
    if metadata_entries:
        item = metadata_entries[0]
        metadata_path = downloads / item["key"]
        print(f"Downloading {item['key']} ...")
        download_file(session, item["url"], metadata_path, timeout, force)

    if force and extracted.exists():
        shutil.rmtree(extracted)
    extracted.mkdir(parents=True, exist_ok=True)

    for archive in downloaded_zips:
        marker = extracted / f".extracted_{archive.stem}"
        if marker.exists() and not force:
            continue

        with zipfile.ZipFile(archive, "r") as zf:
            for member in zf.infolist():
                if member.is_dir():
                    continue

                member_path = Path(member.filename)
                if member_path.is_absolute() or ".." in member_path.parts:
                    raise RuntimeError(
                        f"Unsafe ZIP member in {archive.name}: {member.filename}"
                    )

                target = extracted / member_path
                target.parent.mkdir(parents=True, exist_ok=True)
                with zf.open(member) as src, target.open("wb") as dst:
                    shutil.copyfileobj(src, dst)

        marker.write_text(
            f"{archive.name}\nsha256={sha256_file(archive)}\n",
            encoding="utf-8",
        )

    provenance = {
        "record_id": record.get("id", ZENODO_RECORD_ID),
        "doi": (
            (record.get("pids") or {}).get("doi", {}).get("identifier")
            or JSMEL_DOI
        ),
        "title": (record.get("metadata") or {}).get("title"),
        "created": record.get("created"),
        "updated": record.get("updated"),
        "landing_page": ZENODO_LANDING,
        "api": ZENODO_RECORD_API,
        "downloaded_files": [
            {
                "name": p.name,
                "sha256": sha256_file(p),
                "size": p.stat().st_size,
            }
            for p in downloaded_zips
        ],
    }

    if metadata_path is not None and metadata_path.exists():
        provenance["metadata_file"] = {
            "name": metadata_path.name,
            "sha256": sha256_file(metadata_path),
            "size": metadata_path.stat().st_size,
        }

    return extracted, metadata_path, provenance


def expand_local_source(source_dir: Path, output: Path, force: bool) -> tuple[Path, Path | None]:
    """
    Accept either already-extracted JSMel files or a directory containing the
    official ZIPs. Archives are extracted under output/_jsmel_source.
    """
    source_dir = source_dir.resolve()

    if list(source_dir.rglob("P_*.txt")):
        metadata = next(
            iter(source_dir.rglob("Metadata_JSMel.xlsx")),
            None,
        )
        return source_dir, metadata

    archives = sorted(source_dir.rglob("*.zip"))
    if not archives:
        raise RuntimeError(
            f"No P_*.txt files or ZIP archives found under {source_dir}"
        )

    extracted = output / "_jsmel_source"
    if force and extracted.exists():
        shutil.rmtree(extracted)
    extracted.mkdir(parents=True, exist_ok=True)

    for archive in archives:
        with zipfile.ZipFile(archive, "r") as zf:
            for member in zf.infolist():
                if member.is_dir():
                    continue
                member_path = Path(member.filename)
                if member_path.is_absolute() or ".." in member_path.parts:
                    raise RuntimeError(
                        f"Unsafe ZIP member in {archive.name}: {member.filename}"
                    )
                target = extracted / member_path
                target.parent.mkdir(parents=True, exist_ok=True)
                if target.exists() and not force:
                    continue
                with zf.open(member) as src, target.open("wb") as dst:
                    shutil.copyfileobj(src, dst)

    metadata = next(
        iter(source_dir.rglob("Metadata_JSMel.xlsx")),
        None,
    )
    return extracted, metadata


def melody_id_from_path(path: Path, prefix: str) -> str:
    stem = path.stem
    if not stem.upper().startswith(prefix.upper()):
        raise ValueError(f"Unexpected sequence filename: {path.name}")
    return stem[len(prefix):].strip()


def find_pairs(root: Path) -> tuple[list[Pair], list[str]]:
    pitches: dict[str, Path] = {}
    durations: dict[str, Path] = {}

    for path in root.rglob("*.txt"):
        name = path.stem
        upper = name.upper()
        if upper.startswith("P_"):
            pitches[melody_id_from_path(path, "P_").upper()] = path
        elif upper.startswith("D_"):
            durations[melody_id_from_path(path, "D_").upper()] = path

    ids = sorted(set(pitches) | set(durations))
    pairs: list[Pair] = []
    failures: list[str] = []

    for melody_id in ids:
        p = pitches.get(melody_id)
        d = durations.get(melody_id)
        if p is None or d is None:
            failures.append(
                f"{melody_id}: missing "
                + ("pitch file" if p is None else "duration file")
            )
            continue
        pairs.append(Pair(melody_id, p, d))

    return pairs, failures


def parse_pitch_tokens(text: str) -> list[str]:
    """
    Robust to one-token-per-line, whitespace, comma-separated, bracketed,
    quoted, or MATLAB-ish text representations.
    """
    text = text.replace("\ufeff", "")
    pattern = re.compile(
        r"(?i)(?:rest)|(?:[A-G](?:#{1,2}|b{1,2}|x|♯|♭)?-?\d+)"
    )
    return [m.group(0) for m in pattern.finditer(text)]


def parse_duration_token(token: str) -> float:
    token = token.strip()
    if "/" in token:
        a, b = token.split("/", 1)
        value = float(a) / float(b)
    else:
        value = float(token)

    if not math.isfinite(value) or value <= 0:
        raise ValueError(f"Invalid duration value: {token!r}")
    return value


def parse_duration_tokens(text: str) -> list[float]:
    text = text.replace("\ufeff", "")
    return [
        parse_duration_token(m.group(0))
        for m in NUMBER_RE.finditer(text)
    ]


def pitch_to_midi(token: str) -> int | None:
    if token.strip().lower() == "rest":
        return None

    normalized = (
        token.strip()
        .replace("♯", "#")
        .replace("♭", "b")
    )
    m = PITCH_RE.match(normalized)
    if not m:
        raise ValueError(f"Unsupported JSMel pitch token: {token!r}")

    name = m.group("name").upper()
    acc = m.group("acc")
    octave = int(m.group("oct"))

    pc = {
        "C": 0, "D": 2, "E": 4, "F": 5,
        "G": 7, "A": 9, "B": 11,
    }[name]

    if acc == "#":
        pc += 1
    elif acc == "##" or acc == "x":
        pc += 2
    elif acc == "b":
        pc -= 1
    elif acc == "bb":
        pc -= 2
    elif acc:
        raise ValueError(f"Unsupported accidental: {acc!r}")

    midi = 12 * (octave + 1) + pc
    if not 0 <= midi <= 127:
        raise ValueError(f"MIDI pitch out of range for {token!r}: {midi}")

    return int(midi)


def load_events(pair: Pair) -> list[Event]:
    pitch_text = pair.pitch_path.read_text(
        encoding="utf-8-sig", errors="replace"
    )
    duration_text = pair.duration_path.read_text(
        encoding="utf-8-sig", errors="replace"
    )

    pitch_tokens = parse_pitch_tokens(pitch_text)
    durations = parse_duration_tokens(duration_text)

    if not pitch_tokens:
        raise ValueError(f"No pitch tokens in {pair.pitch_path}")
    if not durations:
        raise ValueError(f"No duration tokens in {pair.duration_path}")
    if len(pitch_tokens) != len(durations):
        raise ValueError(
            f"{pair.melody_id}: pitch/duration length mismatch "
            f"({len(pitch_tokens)} != {len(durations)})"
        )

    return [
        Event(pitch=pitch_to_midi(p), duration_q=float(d))
        for p, d in zip(pitch_tokens, durations)
    ]


def midi_to_hz(midi: int) -> float:
    return 440.0 * 2.0 ** ((midi - 69) / 12.0)


def neutral_tone(freq: float, seconds: float, sr: int) -> np.ndarray:
    n = max(1, int(round(seconds * sr)))
    t = np.arange(n, dtype=np.float64) / sr
    phase = 2.0 * np.pi * freq * t

    # Same deliberately neutral reed-like harmonic recipe used by the earlier
    # Scaleify country corpus generators.
    y = (
        0.78 * np.sin(phase)
        + 0.15 * np.sin(2.0 * phase)
        + 0.05 * np.sin(3.0 * phase)
        + 0.02 * np.sin(4.0 * phase)
    )

    attack = min(n, max(1, int(round(0.008 * sr))))
    release = min(n, max(1, int(round(0.030 * sr))))

    env = np.ones(n, dtype=np.float64)
    if attack > 1:
        env[:attack] = np.linspace(0.0, 1.0, attack, endpoint=False)
    if release > 1:
        env[-release:] *= np.linspace(1.0, 0.0, release)

    return (0.78 * y * env).astype(np.float32)


def render_events(
    events: list[Event],
    sr: int,
    bpm: float,
    gap_ms: float,
) -> np.ndarray:
    q_seconds = 60.0 / bpm
    total_seconds = sum(e.duration_q for e in events) * q_seconds
    total_samples = max(1, int(math.ceil(total_seconds * sr)))
    audio = np.zeros(total_samples, dtype=np.float32)

    cursor_seconds = 0.0
    gap_seconds = max(0.0, gap_ms / 1000.0)

    for event in events:
        duration_seconds = event.duration_q * q_seconds

        if event.pitch is not None:
            # Preserve the symbolic duration exactly, but leave a tiny silence
            # at the end of a sounded note to help the current audio trainer
            # distinguish repeated identical pitches.
            gap = min(gap_seconds, duration_seconds * 0.12)
            sounded = max(0.018, duration_seconds - gap)

            tone = neutral_tone(
                midi_to_hz(event.pitch),
                sounded,
                sr,
            )
            start = int(round(cursor_seconds * sr))
            end = min(total_samples, start + len(tone))
            if end > start:
                audio[start:end] += tone[: end - start]

        cursor_seconds += duration_seconds

    peak = float(np.max(np.abs(audio)))
    if peak > 0:
        audio *= 0.88 / peak

    return audio


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value)).strip().lower()
    text = re.sub(r"[\s_\-/()（）\[\]:：]+", "", text)
    return text


def read_metadata(path: Path | None) -> tuple[dict[str, dict], list[str]]:
    """
    Read metadata without assuming an exact worksheet/header row. The paper
    specifies the semantic fields, but repository workbook formatting may
    evolve between versions.
    """
    if path is None or not path.exists():
        return {}, []

    wb = load_workbook(path, read_only=True, data_only=True)
    records: dict[str, dict] = {}
    diagnostics: list[str] = []

    id_aliases = {
        "melodyid", "id", "melodyidentifier", "楽曲id", "曲id",
    }

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_idx = None
        id_col = None

        # Search first 25 rows for a header containing Melody ID.
        for idx, row in enumerate(rows[:25]):
            normalized = [normalize_header(x) for x in row]
            for j, h in enumerate(normalized):
                if h in id_aliases or ("melody" in h and "id" in h):
                    header_idx = idx
                    id_col = j
                    break
            if header_idx is not None:
                break

        if header_idx is None or id_col is None:
            continue

        headers_raw = rows[header_idx]
        headers = [
            str(h).strip() if h is not None else f"column_{i+1}"
            for i, h in enumerate(headers_raw)
        ]

        for row in rows[header_idx + 1:]:
            if id_col >= len(row):
                continue
            raw_id = row[id_col]
            if raw_id is None:
                continue

            melody_id = str(raw_id).strip().upper()
            if not re.match(r"^[A-R]\d{3}$", melody_id):
                continue

            record = {}
            for i, header in enumerate(headers):
                if i < len(row):
                    value = row[i]
                    if value is not None:
                        record[header] = value

            record["_worksheet"] = ws.title
            records[melody_id] = record

    if not records:
        diagnostics.append(
            "Metadata workbook was found, but no Melody ID column could be parsed."
        )

    return records, diagnostics


def metadata_lookup(record: dict, keywords: Iterable[str]) -> object | None:
    normalized_keywords = [normalize_header(k) for k in keywords]

    for key, value in record.items():
        nk = normalize_header(key)
        if all(k in nk for k in normalized_keywords):
            return value
    return None


def foreign_origin_value(record: dict) -> object | None:
    # Paper's semantic field: "Songs of foreign origin".
    value = metadata_lookup(record, ("foreign", "origin"))
    if value is not None:
        return value

    # Japanese workbook fallback.
    for key, value in record.items():
        nk = normalize_header(key)
        if "外国" in nk and ("由来" in nk or "原曲" in nk or "起源" in nk):
            return value
    return None


def is_foreign_origin(record: dict) -> bool | None:
    if not record:
        return None

    value = foreign_origin_value(record)
    if value is None:
        return None

    text = unicodedata.normalize("NFKC", str(value)).strip().lower()
    if text in FALSEISH:
        return False

    if text in {"yes", "y", "true", "1", "○", "〇", "あり", "有"}:
        return True

    # If this field contains an explicit country/origin name rather than a
    # boolean marker, treat any non-Japan non-empty value as foreign.
    if "japan" in text or "日本" in text:
        return False
    return bool(text)


def compact_metadata(record: dict) -> dict:
    if not record:
        return {}

    out = {"worksheet": record.get("_worksheet")}

    semantic_fields = {
        "song_title": [("song", "title")],
        "composer": [("composer",)],
        "copyright_status": [("copyright", "status")],
        "foreign_origin": [("foreign", "origin")],
        "book_category": [("book", "category")],
        "number_of_notes": [("number", "notes")],
        "number_of_measures": [("number", "measures")],
    }

    for target, candidates in semantic_fields.items():
        for keywords in candidates:
            value = metadata_lookup(record, keywords)
            if value is not None:
                out[target] = value
                break

    # Keep the raw row too, because column names may vary by repository version.
    out["raw"] = {
        str(k): v
        for k, v in record.items()
        if not str(k).startswith("_")
    }
    return out


def write_csv(path: Path, rows: list[dict]) -> None:
    fields = [
        "filename",
        "melody_id",
        "events",
        "notes",
        "rests",
        "duration_quarters",
        "duration_seconds",
        "pitch_min_midi",
        "pitch_max_midi",
        "foreign_origin",
        "source_pitch_file",
        "source_duration_file",
    ]

    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in fields})


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Generate neutral WAVs from the 547-song open public-domain "
            "subset of JSMel."
        )
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Default: dataset/japan_jsmel",
    )
    parser.add_argument(
        "--source-dir",
        type=Path,
        default=None,
        help=(
            "Use already-downloaded/extracted JSMel files instead of fetching "
            "the current open Zenodo record."
        ),
    )
    parser.add_argument(
        "--song",
        type=str,
        default=None,
        help="Generate one melody ID, e.g. A001.",
    )
    parser.add_argument(
        "--exclude-foreign-origin",
        action="store_true",
        help=(
            "Skip melodies marked as foreign-origin in Metadata_JSMel.xlsx. "
            "Rows whose foreign-origin status cannot be parsed are retained."
        ),
    )
    parser.add_argument("--list", action="store_true")
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--sample-rate", type=int, default=DEFAULT_SR)
    parser.add_argument("--render-bpm", type=float, default=DEFAULT_BPM)
    parser.add_argument(
        "--articulation-gap-ms",
        type=float,
        default=DEFAULT_GAP_MS,
    )
    parser.add_argument("--timeout", type=float, default=60.0)
    parser.add_argument("--request-delay", type=float, default=0.10)
    args = parser.parse_args()

    if args.sample_rate < 8000:
        parser.error("--sample-rate must be >= 8000")
    if args.render_bpm <= 0:
        parser.error("--render-bpm must be > 0")
    if args.articulation_gap_ms < 0:
        parser.error("--articulation-gap-ms must be >= 0")

    output = args.output
    output.mkdir(parents=True, exist_ok=True)

    if args.source_dir is not None:
        source_root, metadata_path = expand_local_source(
            args.source_dir,
            output,
            args.force,
        )
        provenance = {
            "mode": "local_source",
            "source_dir": str(args.source_dir.resolve()),
            "jsmel_doi": JSMEL_DOI,
        }
    else:
        source_root, metadata_path, provenance = acquire_jsmel(
            output=output,
            timeout=args.timeout,
            force=args.force,
            request_delay=args.request_delay,
        )
        provenance["mode"] = "zenodo_download"

    pairs, pairing_failures = find_pairs(source_root)
    if not pairs:
        raise RuntimeError(
            f"No JSMel P_/D_ sequence pairs found under {source_root}"
        )

    metadata, metadata_diagnostics = read_metadata(metadata_path)

    if args.song:
        target = args.song.strip().upper()
        pairs = [p for p in pairs if p.melody_id.upper() == target]
        if not pairs:
            raise RuntimeError(f"Melody ID not found in open PD files: {target}")

    if args.exclude_foreign_origin:
        pairs = [
            p for p in pairs
            if is_foreign_origin(metadata.get(p.melody_id, {})) is not True
        ]

    if args.list:
        print(
            f"{'ID':8} {'FOREIGN':8} {'PITCH FILE':32} DURATION FILE"
        )
        print("-" * 100)
        for pair in pairs:
            foreign = is_foreign_origin(metadata.get(pair.melody_id, {}))
            label = "yes" if foreign is True else "no" if foreign is False else "?"
            print(
                f"{pair.melody_id:8} {label:8} "
                f"{pair.pitch_path.name:32} {pair.duration_path.name}"
            )
        print()
        print(f"Detected usable PD pairs: {len(pairs)}")
        return

    rows: list[dict] = []
    failures: list[dict] = []

    print(f"Detected sequence pairs: {len(pairs)}")
    if metadata:
        print(f"Metadata rows parsed:    {len(metadata)}")
    print()

    for i, pair in enumerate(pairs, start=1):
        print(f"[{i:03d}/{len(pairs):03d}] {pair.melody_id}")
        wav_path = output / f"{pair.melody_id}.wav"

        try:
            events = load_events(pair)
            notes = [e.pitch for e in events if e.pitch is not None]
            rests = sum(1 for e in events if e.pitch is None)
            duration_q = float(sum(e.duration_q for e in events))

            if not notes:
                raise ValueError("Melody contains no pitched events")

            if args.force or not wav_path.exists():
                audio = render_events(
                    events=events,
                    sr=args.sample_rate,
                    bpm=args.render_bpm,
                    gap_ms=args.articulation_gap_ms,
                )
                sf.write(
                    wav_path,
                    audio,
                    args.sample_rate,
                    subtype="PCM_16",
                )

            foreign = is_foreign_origin(
                metadata.get(pair.melody_id, {})
            )

            row = {
                "filename": wav_path.name,
                "melody_id": pair.melody_id,
                "events": len(events),
                "notes": len(notes),
                "rests": rests,
                "duration_quarters": round(duration_q, 6),
                "duration_seconds": round(
                    duration_q * 60.0 / args.render_bpm,
                    6,
                ),
                "pitch_min_midi": min(notes),
                "pitch_max_midi": max(notes),
                "foreign_origin": (
                    "yes" if foreign is True
                    else "no" if foreign is False
                    else "unknown"
                ),
                "source_pitch_file": str(
                    pair.pitch_path.relative_to(source_root)
                ),
                "source_duration_file": str(
                    pair.duration_path.relative_to(source_root)
                ),
                "metadata": compact_metadata(
                    metadata.get(pair.melody_id, {})
                ),
            }
            rows.append(row)
            print(
                f"    events={len(events)} notes={len(notes)} rests={rests} "
                f"range={min(notes)}-{max(notes)} "
                f"duration={row['duration_seconds']:.2f}s"
            )

        except Exception as exc:
            failures.append({
                "melody_id": pair.melody_id,
                "pitch_file": str(pair.pitch_path),
                "duration_file": str(pair.duration_path),
                "error": f"{type(exc).__name__}: {exc}",
            })
            print(f"    [FAILED] {type(exc).__name__}: {exc}")

    write_csv(output / "metadata.csv", rows)

    manifest = {
        "dataset": "JSMel public-domain subset",
        "dataset_version_source": provenance,
        "paper_doi": PAPER_DOI,
        "open_dataset_doi": JSMEL_DOI,
        "license": "CC BY-NC 4.0 (JSMel open symbolic data and metadata)",
        "scope": (
            "Melodies widely shared in contemporary Japan; not restricted to "
            "Japanese-origin traditional repertoire."
        ),
        "public_domain_subset_size_reported_by_jsmel": 547,
        "sequence_pairs_detected_before_filters": (
            len(find_pairs(source_root)[0])
        ),
        "exclude_foreign_origin": bool(args.exclude_foreign_origin),
        "rendering": {
            "sample_rate": args.sample_rate,
            "format": "WAV PCM16 mono",
            "render_bpm": args.render_bpm,
            "quarter_note_seconds": 60.0 / args.render_bpm,
            "articulation_gap_ms": args.articulation_gap_ms,
            "timbre": "Scaleify neutral reed-like synth",
            "pitch_estimation_used": False,
            "source_accompaniment_used": False,
        },
        "generated_count": len(rows),
        "failed_count": len(failures),
        "pairing_diagnostics": pairing_failures,
        "metadata_diagnostics": metadata_diagnostics,
        "songs": rows,
        "failures": failures,
        "citation": (
            "Matsunaga R, Ishimoto T, Hartono P, Abe J. "
            "JSMel: A dataset of 960 song melodies widely shared in "
            "contemporary Japan. Data in Brief. 2026;66:112869."
        ),
    }

    (output / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, default=str) + "\n",
        encoding="utf-8",
    )

    if failures:
        (output / "failures.json").write_text(
            json.dumps(failures, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )

    print()
    print("Generation complete")
    print("-------------------")
    print(f"Generated: {len(rows)}")
    print(f"Failed:    {len(failures)}")
    print(f"Output:    {output}")
    print(f"Metadata:  {output / 'metadata.csv'}")
    print(f"Manifest:  {output / 'manifest.json'}")

    if failures:
        raise SystemExit(2)


if __name__ == "__main__":
    main()